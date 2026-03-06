"""
SAI Rendicontazione — App Streamlit
====================================
Carica il registro Excel + il PDF mensile dei bonifici Intesa Sanpaolo.
L'app:
  1. Estrae ogni bonifico dal PDF (una pagina = un bonifico)
  2. Abbina ogni bonifico a una o più righe del registro tramite:
       - Numero fattura (causale PDF ↔ N_Documento registro)  [metodo primario]
       - Importo esatto (solo se 1 riga corrisponde)           [fallback]
       - Nome beneficiario fuzzy (per stipendi/locazioni)      [fallback 2]
  3. Genera un nome file per ciascun PDF nel formato:
       {N°Reg} - BONIFICO - {Beneficiario} - FATT. N. {Fatt} - {DataFatt}.pdf
  4. Produce uno ZIP con i PDF rinominati (3 cartelle: abbinati / da_verificare / interni)
  5. Produce il registro Excel con Modalità pagamento + Data pagamento compilati
     (celle evidenziate in verde; non sovrascrive campi già presenti)

Avvio:
    pip install streamlit pdfplumber pandas openpyxl pypdf
    streamlit run sai_rendicontazione.py
"""

# ── imports ──────────────────────────────────────────────────────────────────
import io
import re
import zipfile
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
import pdfplumber
import pypdf
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURAZIONE PAGINA
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="SAI · Rendicontazione",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #F5F4F0; }
.block-container { padding: 2rem 3rem; max-width: 1400px; }
h1, h2, h3 { font-family: 'DM Sans', sans-serif; font-weight: 600; }

/* ── header ── */
.hdr {
    background: #1C1C1C; border-radius: 14px;
    padding: 1.75rem 2.25rem; margin-bottom: 2rem;
}
.hdr h1 { color: #FFF; font-size: 1.5rem; margin: 0 0 .25rem; letter-spacing: -.02em; }
.hdr p  { color: #888; font-size: .82rem; margin: 0; font-family: 'DM Mono', monospace; }
.badge  {
    background: #C5F135; color: #1C1C1C;
    font-size: .62rem; font-family: 'DM Mono', monospace; font-weight: 500;
    padding: .2rem .5rem; border-radius: 4px; letter-spacing: .06em;
    vertical-align: middle; margin-left: .4rem;
}

/* ── metriche ── */
.mrow { display: flex; gap: .9rem; margin-bottom: 1.5rem; flex-wrap: wrap; }
.mbox {
    flex: 1; min-width: 110px;
    background: white; border: 1px solid #E5E5E1;
    border-radius: 10px; padding: 1rem 1.25rem;
}
.mlbl { font-size: .65rem; font-family: 'DM Mono', monospace; text-transform: uppercase;
        letter-spacing: .07em; color: #999; margin-bottom: .2rem; }
.mval { font-size: 1.8rem; font-weight: 600; color: #1C1C1C; line-height: 1; }
.msub { font-size: .72rem; color: #AAA; margin-top: .1rem; }

/* ── tabella ── */
.th {
    display: grid;
    grid-template-columns: 40px 90px 1fr 120px 100px 90px;
    gap: .4rem; padding: .55rem .9rem;
    background: #EFEFEB; border-radius: 8px 8px 0 0;
    border: 1px solid #E5E5E1;
    font-family: 'DM Mono', monospace; font-size: .62rem;
    text-transform: uppercase; letter-spacing: .07em; color: #999;
}
.tb {
    border: 1px solid #E5E5E1; border-top: none;
    border-radius: 0 0 8px 8px; background: white;
    overflow: hidden; margin-bottom: 1.5rem;
}
.tr {
    display: grid;
    grid-template-columns: 40px 90px 1fr 120px 100px 90px;
    gap: .4rem; padding: .6rem .9rem;
    border-bottom: 1px solid #F3F3F0;
    align-items: start; font-size: .8rem;
}
.tr:last-child { border-bottom: none; }
.tr:hover { background: #FAFAF8; }

.fn {
    font-family: 'DM Mono', monospace; font-size: .68rem;
    background: #F5F4F0; border: 1px solid #E0E0DA;
    border-radius: 5px; padding: .22rem .45rem;
    color: #444; word-break: break-all; line-height: 1.4;
}
.mono { font-family: 'DM Mono', monospace; font-size: .75rem; color: #555; }
.av   { font-size: .68rem; color: #B85C00; margin-top: .2rem; }

.bdg { display: inline-block; padding: .18rem .55rem; border-radius: 20px;
       font-size: .65rem; font-family: 'DM Mono', monospace; font-weight: 500; }
.ba  { background: #E6F9D8; color: #2E7D0C; }
.bm  { background: #FFF5D0; color: #956800; }
.bb  { background: #FFE8CC; color: #B85C00; }
.be  { background: #FDECEA; color: #B71C1C; }
.bi  { background: #EFEFEB; color: #888; }

/* ── bottoni ── */
div.stButton > button {
    background: #1C1C1C !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    padding: .65rem 1.5rem !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important; font-size: .95rem !important; width: 100%;
}
.stDownloadButton button {
    background: #C5F135 !important; color: #1C1C1C !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; width: 100%;
}
.stProgress > div > div { background: #C5F135 !important; }
hr { border: none; border-top: 1px solid #E5E5E1; margin: 1.5rem 0; }
.slbl { font-size: .65rem; font-family: 'DM Mono', monospace;
        text-transform: uppercase; letter-spacing: .08em; color: #999; margin-bottom: .35rem; }
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
#  FUNZIONI CORE
# ─────────────────────────────────────────────────────────────────────────────

def sanitize(s: str) -> str:
    """Slash -> trattino (3/757 -> 3-757) per nomi file leggibili."""
    s = str(s).strip()
    s = s.replace("/", "-")
    s = re.sub(r'[\\*?:"<>|]', "", s)
    s = re.sub(r"\s+", " ", s)
    return s[:80]


def parse_importo(s: str) -> float | None:
    try:
        return float(str(s).replace(".", "").replace(",", ".").strip())
    except Exception:
        return None


def token_sim(a: str, b: str) -> int:
    """Similarità 0-100 basata su token condivisi (no librerie esterne)."""
    a_tok = set(re.split(r"[\s\-_/]+", a.upper())) - {""}
    b_tok = set(re.split(r"[\s\-_/]+", b.upper())) - {""}
    if not a_tok or not b_tok:
        return 0
    return int(100 * len(a_tok & b_tok) / max(len(a_tok), len(b_tok)))


# ── estrazione PDF ────────────────────────────────────────────────────────────

def _numeri_fattura(causale: str) -> list[str]:
    """
    Estrae i numeri di fattura dalla causale.
    Casi gestiti:
      "nr. 3/757"                      → ["3/757"]
      "nr. 449/2025"                   → ["449/2025"]
      "nr. 1/25/W"                     → ["1/25/W"]
      "nr. FPR 246/25"                 → ["FPR 246/25"]
      "nr. 1107/B … nr. 1109/B …"     → ["1107/B", "1109/B", "1108/B"]
      "nr. 3752/3818/4461/4427/4338/…" → lista compatta ≥5 slash
    """
    # 1. Lista compatta con ≥ 5 slash (es. 3752/3818/4461/4427/4338/4444)
    m = re.search(r"nr\.?\s+([\w]+(?:/[\w]+){4,})", causale, re.I)
    if m:
        return [p for p in m.group(1).split("/") if p.strip()]

    # 2. Prefisso FPR speciale: "nr. FPR 246/25"
    m = re.search(r"nr\.?\s+(FPR\s+[\w/]+)", causale, re.I)
    if m:
        return [m.group(1).strip()]

    # 3. Tutti i "nr. X" con lookahead per fermarsi a "del", "nr.", fine riga
    raw = re.findall(
        r"nr\.?\s+([\w/.\-]+?)(?=\s+del\b|\s+nr\.?\b|\s+e\s+ricevute|\s*$|\s*-\s)",
        causale, re.I
    )
    if not raw:
        raw = re.findall(r"nr\.?\s+([\w/.\-]+)", causale, re.I)

    out = []
    for n in raw:
        n = n.strip().rstrip(".-")
        if not n:
            continue
        if re.match(r"\d{2}/\d{2}/\d{4}", n):  # data GG/MM/AAAA
            continue
        if n.lower() in ("del", "e", "e.", "al"):
            continue
        out.append(n)
    return out


def estrai_bonifici(pdf_bytes: bytes) -> list[dict]:
    """Una pagina = un bonifico Intesa Sanpaolo."""
    out = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""

            # Beneficiario
            m = re.search(r"Beneficiario:\s*([^\n]+)", text)
            ben = m.group(1).strip() if m else ""
            ben = re.split(r"\s*-\s*(?:LEI|Persona)", ben)[0].strip().split("\n")[0].strip()

            # Importo
            m = re.search(r"Importo da trasferire:\s*([\d\.,]+)\s*EUR", text)
            imp_str = m.group(1) if m else ""
            imp_val = parse_importo(imp_str) if imp_str else None

            # Date
            m = re.search(r"Data di addebito:\s*(\d{2}\.\d{2}\.\d{4})", text)
            data_add = m.group(1) if m else ""
            m = re.search(r"Data creazione:\s*(\d{2}\.\d{2}\.\d{4})", text)
            data_cre = m.group(1) if m else ""
            data_pag = data_add or data_cre

            # Causale
            m = re.search(r"Informazioni aggiuntive \(max\s+(.+)", text)
            causale = ""
            if m:
                causale = re.sub(r"^140 caratteri\)\s*", "", m.group(1).strip()).strip()

            numeri = _numeri_fattura(causale)

            # Prima data "del GG/MM/AAAA" nella causale
            m = re.search(r"\bdel\s+(\d{2}/\d{2}/\d{4})", causale, re.I)
            data_fatt = m.group(1) if m else ""

            out.append(
                dict(
                    pagina=i + 1,
                    beneficiario=ben,
                    importo=imp_val,
                    importo_str=imp_str,
                    data_pagamento=data_pag,
                    causale=causale,
                    numeri_fattura=numeri,
                    data_fattura=data_fatt,
                )
            )
    return out


# ── caricamento registro ──────────────────────────────────────────────────────

def carica_registro(excel_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(excel_bytes), header=6, skiprows=[7])
    df.columns = [
        "_", "N", "Natura", "Data_Doc", "N_Documento",
        "Modalita_Pagamento", "Data_Pagamento", "Cod_Spesa",
        "Descrizione", "Importo_Totale", "Finanziamento",
        "Importo_Imputato", "Coop",
    ]
    df = df[df["N"].notna() & (df["N"] != "N.")].copy()
    df["N"]            = df["N"].astype(str).str.strip()
    df["N_Documento"]  = df["N_Documento"].astype(str).str.strip()
    df["Desc_norm"]    = df["Descrizione"].astype(str).str.upper().str.strip()
    df["Importo_Totale"] = pd.to_numeric(df["Importo_Totale"], errors="coerce")
    return df


# ── abbinamento ───────────────────────────────────────────────────────────────

_PAT_INTERNO = [
    r"ricarica\s+(cassa|fondo)",
    r"giroconto",
    r"pocket\s+money",
    r"\bvitto\s+(gen|feb|mar|apr|mag|giu|lug|ago|set|ott|nov|dic)",
]
_BEN_INTERNO = ["carta prepagata", "i girasoli scs"]


def is_interno(b: dict) -> bool:
    causale = b["causale"].lower()
    ben     = b["beneficiario"].lower()
    if any(p in ben for p in _BEN_INTERNO):
        return True
    for pat in _PAT_INTERNO:
        if re.search(pat, causale):
            return True
    return False


def _ndoc_match(ndoc: str, nf: str, nf_norm: str) -> bool:
    nd_norm = re.sub(r"^FPR\s*", "", ndoc, flags=re.I).strip()
    return ndoc == nf or ndoc == nf_norm or nd_norm == nf_norm


def abbina(bon: dict, reg: pd.DataFrame) -> tuple[list[str], str, str]:
    """
    Ritorna (n_registro_list, metodo, confidence).
    confidence: "alta" | "media" | "bassa" | "nessuna"
    """
    mask_b = reg["Modalita_Pagamento"].astype(str).str.upper().str.contains("BONI", na=False)
    mask_v = reg["Modalita_Pagamento"].isna() | reg["Modalita_Pagamento"].astype(str).str.strip().isin(["", "nan", "None"])
    sub = reg[mask_b | mask_v].copy()

    # 1 ── Numero fattura ─────────────────────────────────────────────────
    if bon["numeri_fattura"]:
        trovati: list[str] = []
        for nf in bon["numeri_fattura"]:
            nf = nf.strip()
            nf_norm = re.sub(r"^FPR\s*", "", nf, flags=re.I).strip()
            hit = sub[sub["N_Documento"].apply(lambda x: _ndoc_match(str(x).strip(), nf, nf_norm))]
            trovati.extend(hit["N"].tolist())
        # dedup preservando ordine
        seen: set = set()
        trovati = [n for n in trovati if not (n in seen or seen.add(n))]  # type: ignore[func-returns-value]
        if trovati:
            somma = reg[reg["N"].isin(trovati)]["Importo_Totale"].sum()
            diff  = abs(somma - (bon["importo"] or 0))
            return trovati, "fattura", "alta" if diff < 1.0 else "media"

    # 2 ── Importo esatto (1 sola riga) ──────────────────────────────────
    if bon["importo"]:
        fi = sub[abs(sub["Importo_Totale"] - bon["importo"]) < 0.02]
        if len(fi) == 1:
            return fi["N"].tolist(), "importo", "media"

    # 3 ── Nome fuzzy (stipendi / locazioni) ─────────────────────────────
    ben_c = re.sub(r"\b(NUOVO|NEW|CONTO|SRL|SRLS|SPA|SNC|SAS)\b", "", bon["beneficiario"].upper()).strip()
    best_s, best_n = 0, None
    for _, row in sub.iterrows():
        s = token_sim(ben_c, str(row["Desc_norm"]))
        if s > best_s:
            best_s, best_n = s, row["N"]

    if best_s >= 75 and best_n:
        imp_ok = True
        if bon["importo"]:
            r = reg[reg["N"] == best_n]
            if not r.empty and abs(r.iloc[0]["Importo_Totale"] - bon["importo"]) > 5:
                imp_ok = False
        conf = ("alta" if best_s >= 90 else "media") if imp_ok else "bassa"
        return [best_n], "nome_fuzzy", conf

    return [], "nessuno", "nessuna"


# ── nome file ─────────────────────────────────────────────────────────────────

def build_nome(n_reg: str, ben: str, numeri: list[str], data_fatt: str) -> str:
    """
    {N°Reg} - BONIFICO - {Beneficiario} - FATT. N. {xxx} - {DD-MM-YYYY}.pdf
    La data viene omessa se assente.
    """
    parti = [
        sanitize(n_reg) if n_reg else "???",
        "BONIFICO",
        sanitize(ben),
        f"FATT. N. {sanitize(' + '.join(numeri))}" if numeri else "FATT. N. -",
    ]
    if data_fatt:
        parti.append(data_fatt.replace("/", "-").replace(".", "-"))
    return " - ".join(parti) + ".pdf"


# ── estrazione singola pagina PDF ─────────────────────────────────────────────

def estrai_pagina(pdf_bytes: bytes, idx: int) -> bytes | None:
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        writer = pypdf.PdfWriter()
        writer.add_page(reader.pages[idx])
        buf = io.BytesIO()
        writer.write(buf)
        return buf.getvalue()
    except Exception:
        return None


# ── compilazione registro ─────────────────────────────────────────────────────

def compila_registro(
    excel_bytes: bytes, risultati: list[dict], reg_df: pd.DataFrame
) -> tuple[bytes, int]:
    """
    Compila N_Documento (se mancante), Modalita_Pagamento e Data_Pagamento
    per le righe abbinate con confidence alta/media/bassa.
    Celle aggiornate → verde chiaro. Celle già compilate → non toccate.
    """
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    # Struttura fissa del registro:
    # Col B(2)=N, E(5)=N_Documento, F(6)=Modalita_Pagamento, G(7)=Data_Pagamento
    COL_N    = 2
    COL_NDOC = 5
    COL_MOD  = 6
    COL_DATA = 7

    n_to_row: dict[str, int] = {}
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        c = row[COL_N - 1]
        v = str(c.value or "").strip()
        if v and v not in ("N.", "nan"):
            n_to_row[v] = c.row

    verde = PatternFill("solid", fgColor="D4F5C8")
    vfont = Font(color="1A6B0A")
    compilate = 0

    for ris in risultati:
        if ris["confidence"] not in ("alta", "media", "bassa"):
            continue
        if not ris["n_registro"]:
            continue

        dt = None
        dp = ris["data_pagamento"]
        if dp:
            try:
                dt = datetime.strptime(dp, "%d.%m.%Y")
            except ValueError:
                pass

        for n_reg in ris["n_registro"]:
            ex_row = n_to_row.get(str(n_reg))
            if ex_row is None:
                continue

            # Cerca riga nel DataFrame per ottenere N_Documento
            df_riga = reg_df[reg_df["N"] == str(n_reg)]

            mod = False

            # ── N_Documento ──────────────────────────────────────────
            # Se c'è un numero fattura nel bonifico e la cella è vuota
            c_ndoc = ws.cell(row=ex_row, column=COL_NDOC)
            if ris["numeri_fattura"] and str(c_ndoc.value or "").strip() in ("", "nan", "None"):
                # Prendi il numero fattura che matcha questa riga specifica
                if len(ris["numeri_fattura"]) == 1:
                    c_ndoc.value = ris["numeri_fattura"][0]
                else:
                    # Cerca quale numero corrisponde a questa riga
                    for nf in ris["numeri_fattura"]:
                        nf_norm = re.sub(r"^FPR\s*", "", nf, flags=re.I).strip()
                        if not df_riga.empty:
                            ndoc_val = str(df_riga.iloc[0]["N_Documento"]).strip()
                            if _ndoc_match(ndoc_val, nf, nf_norm):
                                c_ndoc.value = nf
                                break
                if c_ndoc.value and str(c_ndoc.value).strip() not in ("", "nan"):
                    c_ndoc.fill = verde
                    c_ndoc.font = vfont
                    mod = True

            # ── Modalità pagamento ───────────────────────────────────
            c_mod = ws.cell(row=ex_row, column=COL_MOD)
            if str(c_mod.value or "").strip() in ("", "nan", "None"):
                c_mod.value = "BONIFICO"
                c_mod.fill  = verde
                c_mod.font  = vfont
                mod = True

            # ── Data pagamento ───────────────────────────────────────
            if dt:
                c_data = ws.cell(row=ex_row, column=COL_DATA)
                if str(c_data.value or "").strip() in ("", "nan", "None", "NaT"):
                    c_data.value         = dt
                    c_data.number_format = "DD/MM/YYYY"
                    c_data.fill          = verde
                    c_data.font          = vfont
                    mod = True

            if mod:
                compilate += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), compilate


# ─────────────────────────────────────────────────────────────────────────────
#  INTERFACCIA
# ─────────────────────────────────────────────────────────────────────────────

st.markdown(
    """
<div class="hdr">
  <h1>SAI · Rendicontazione <span class="badge">BETA</span></h1>
  <p>Abbina i bonifici bancari al registro · Rinomina i PDF · Compila data pagamento e modalità nel registro</p>
</div>
""",
    unsafe_allow_html=True,
)

# Upload
c1, c2 = st.columns(2)
with c1:
    st.markdown('<div class="slbl">📂 Registro spese (.xlsx)</div>', unsafe_allow_html=True)
    reg_file = st.file_uploader("Registro", type=["xlsx"], key="reg", label_visibility="collapsed")
    if reg_file:
        st.success(f"✓ {reg_file.name}")
with c2:
    st.markdown('<div class="slbl">🏦 PDF bonifici banca (Intesa Sanpaolo)</div>', unsafe_allow_html=True)
    pdf_file = st.file_uploader("PDF", type=["pdf"], key="pdf", label_visibility="collapsed")
    if pdf_file:
        st.success(f"✓ {pdf_file.name}  ({pdf_file.size // 1024} KB)")

st.markdown("<hr>", unsafe_allow_html=True)

if reg_file and pdf_file:
    if st.button("▶  Avvia elaborazione", use_container_width=True):

        prog = st.progress(0, text="Lettura PDF…")
        pdf_bytes   = pdf_file.read()
        excel_bytes = reg_file.read()

        with st.spinner("Estrazione bonifici…"):
            bonifici = estrai_bonifici(pdf_bytes)
        prog.progress(20, text="Lettura registro…")

        reg_df = carica_registro(excel_bytes)
        prog.progress(35, text="Abbinamento…")

        risultati: list[dict] = []
        for idx, bon in enumerate(bonifici):
            prog.progress(35 + int(45 * idx / max(len(bonifici), 1)),
                          text=f"Abbinamento {idx + 1}/{len(bonifici)}…")

            if is_interno(bon):
                ris = {
                    **bon,
                    "n_registro": [],
                    "metodo":     "interno",
                    "confidence": "interno",
                    "nome_file":  f"INTERNO - BONIFICO - {sanitize(bon['beneficiario'])}.pdf",
                    "avvisi":     [],
                }
            else:
                n_list, metodo, conf = abbina(bon, reg_df)
                avvisi = []
                if len(n_list) > 1 and metodo == "fattura":
                    somma = reg_df[reg_df["N"].isin(n_list)]["Importo_Totale"].sum()
                    diff  = abs(somma - (bon["importo"] or 0))
                    if diff > 1.0:
                        avvisi.append(f"Somma righe {somma:.2f} € ≠ importo bonifico {bon['importo']:.2f} €")
                ris = {
                    **bon,
                    "n_registro": n_list,
                    "metodo":     metodo,
                    "confidence": conf,
                    "nome_file":  build_nome("+".join(n_list), bon["beneficiario"],
                                             bon["numeri_fattura"], bon["data_fattura"]),
                    "avvisi":     avvisi,
                }
            risultati.append(ris)

        prog.progress(82, text="Compilazione registro…")
        reg_out, n_comp = compila_registro(excel_bytes, risultati, reg_df)

        prog.progress(93, text="Creazione ZIP…")
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for ris in risultati:
                pb = estrai_pagina(pdf_bytes, ris["pagina"] - 1)
                if pb:
                    folder = (
                        "interni"       if ris["confidence"] == "interno"
                        else "da_verificare" if ris["confidence"] in ("bassa", "nessuna")
                        else "abbinati"
                    )
                    zf.writestr(f"{folder}/{ris['nome_file']}", pb)

        prog.progress(100, text="✓ Completato!")
        prog.empty()

        st.session_state.update(
            risultati=risultati, reg_out=reg_out,
            zip_data=zip_buf.getvalue(), n_comp=n_comp, done=True,
        )

# ── Risultati ────────────────────────────────────────────────────────────────
if st.session_state.get("done"):
    R       = st.session_state["risultati"]
    n_comp  = st.session_state["n_comp"]

    tot     = len(R)
    alta    = sum(1 for r in R if r["confidence"] == "alta")
    media   = sum(1 for r in R if r["confidence"] == "media")
    interni = sum(1 for r in R if r["confidence"] == "interno")
    da_ver  = sum(1 for r in R if r["confidence"] in ("bassa", "nessuna"))

    st.markdown(
        f"""
<div class="mrow">
  <div class="mbox"><div class="mlbl">Bonifici PDF</div>
    <div class="mval">{tot}</div></div>
  <div class="mbox"><div class="mlbl">Abbinati — alta</div>
    <div class="mval" style="color:#2E7D0C">{alta}</div></div>
  <div class="mbox"><div class="mlbl">Abbinati — media</div>
    <div class="mval" style="color:#956800">{media}</div></div>
  <div class="mbox"><div class="mlbl">Da verificare</div>
    <div class="mval" style="color:#B71C1C">{da_ver}</div></div>
  <div class="mbox"><div class="mlbl">Interni / esclusi</div>
    <div class="mval" style="color:#888">{interni}</div></div>
  <div class="mbox"><div class="mlbl">Righe compilate</div>
    <div class="mval" style="color:#2E7D0C">{n_comp}</div></div>
</div>
""",
        unsafe_allow_html=True,
    )

    # Tabella
    st.markdown('<div class="slbl">📋 Dettaglio abbinamenti</div>', unsafe_allow_html=True)
    st.markdown(
        """<div class="th">
          <span>Pag.</span><span>Importo</span><span>Nome file PDF generato</span>
          <span>N° Registro</span><span>Metodo</span><span>Stato</span>
        </div><div class="tb">""",
        unsafe_allow_html=True,
    )

    _BADGE = {
        "alta":    '<span class="bdg ba">✓ Alta</span>',
        "media":   '<span class="bdg bm">⚠ Media</span>',
        "bassa":   '<span class="bdg bb">⚠ Bassa</span>',
        "nessuna": '<span class="bdg be">✕ Non trovato</span>',
        "interno": '<span class="bdg bi">↩ Interno</span>',
    }
    _MET = {
        "fattura":    "🔢 N° fatt.",
        "importo":    "💶 Importo",
        "nome_fuzzy": "👤 Fuzzy",
        "interno":    "↩ —",
        "nessuno":    "— —",
    }

    for r in R:
        badge  = _BADGE.get(r["confidence"], "")
        metodo = _MET.get(r["metodo"], r["metodo"])
        n_reg  = ", ".join(r["n_registro"]) if r["n_registro"] else "—"
        imp    = f"€&nbsp;{r['importo_str']}" if r["importo_str"] else "—"
        av_html = "".join(f'<div class="av">⚡ {a}</div>' for a in r.get("avvisi", []))

        st.markdown(
            f"""<div class="tr">
              <span class="mono" style="color:#AAA">{r['pagina']:02d}</span>
              <span class="mono">{imp}</span>
              <span><div class="fn">{r['nome_file']}</div>{av_html}</span>
              <span class="mono">{n_reg}</span>
              <span style="font-size:.75rem;color:#777">{metodo}</span>
              {badge}
            </div>""",
            unsafe_allow_html=True,
        )

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    # Download
    d1, d2 = st.columns(2)
    with d1:
        st.markdown('<div class="slbl">⬇ PDF rinominati (.zip)</div>', unsafe_allow_html=True)
        st.markdown(
            """<p style="font-size:.78rem;color:#777;margin-bottom:.7rem">
            <b>abbinati/</b> — confidence alta o media<br>
            <b>da_verificare/</b> — abbinamento incerto o non trovato<br>
            <b>interni/</b> — pocket money, giroconti, ricariche carta
            </p>""",
            unsafe_allow_html=True,
        )
        st.download_button(
            "⬇  Scarica PDF rinominati (.zip)",
            data=st.session_state["zip_data"],
            file_name="SAI_PDF_rinominati.zip",
            mime="application/zip",
            use_container_width=True,
        )
    with d2:
        st.markdown('<div class="slbl">⬇ Registro Excel aggiornato</div>', unsafe_allow_html=True)
        st.markdown(
            f"""<p style="font-size:.78rem;color:#777;margin-bottom:.7rem">
            Compilate <b>{n_comp} righe</b> nel registro.<br>
            Le celle aggiornate sono in <span style="color:#2E7D0C;font-weight:600">verde</span>.<br>
            I campi già presenti <b>non vengono sovrascritti</b>.
            </p>""",
            unsafe_allow_html=True,
        )
        st.download_button(
            "⬇  Scarica registro aggiornato (.xlsx)",
            data=st.session_state["reg_out"],
            file_name="Registro_aggiornato.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

else:
    if not (reg_file and pdf_file):
        st.markdown(
            """<div style="text-align:center;padding:4rem 2rem;color:#C0C0BC">
              <div style="font-size:3.5rem;margin-bottom:1rem">📂</div>
              <div style="font-family:'DM Mono',monospace;font-size:.88rem">
                Carica il registro Excel e il PDF dei bonifici per iniziare
              </div>
            </div>""",
            unsafe_allow_html=True,
        )
